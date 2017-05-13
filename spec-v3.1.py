import os
from os import walk
import openpyxl
import xlrd
import xlwt
import re


#Рабочий каталог
working_dir = os.getcwd()
#Файл со спецификацией
files = []
final_files = []
for (dirpath, dirnames, filenames) in walk(working_dir):
	files.extend(filenames)
for file in files:
	name, extension = os.path.splitext(file)
	if ((extension == '.xls') | (extension == '.xlsx')) & (name != 'Cisco Product List') & (name != 'Cisco Product List Sorted'):
		final_files.append(file)
i = 0
print ('choose your file:')
for f in final_files:
	print ('[' + str(i) + '] ' + f)
	i += 1

working_spec_filename = final_files[int(input())]
#Итоговый файл
name, extension = os.path.splitext(working_spec_filename)
resulting_file = 'c_' + name + '.xlsx'

def insertion_sort_cpl():
	print ('Sorting CPL')
	cpl_workbook = xlrd.open_workbook('Cisco Product List.xls')
	worksheet = cpl_workbook.sheet_by_name('Cisco Product List Russia')
	array = [worksheet.row_values(i) for i in range(worksheet.nrows)]
	labels = array[0]
	array = array[1:]
	for i in range(1,worksheet.nrows - 1):
		key = array[i]
		j = i - 1
		while (j > 0) & (array[j][0] > key[0]):
			array[j + 1] = array[j]
			j = j - 1
		array[j + 1] = key
	
	bk = xlwt.Workbook()
	sheet = bk.add_sheet(worksheet.name)

	for idx, label in enumerate(labels):
		sheet.write(0, idx, label)
	
	for idx_r, row in enumerate(array):
		for idx_c, value in enumerate(row):
			sheet.write(idx_r+1, idx_c, value)
	
	bk.save('Cisco Product List Sorted.xls')
	print ('CPL Sorted')


def findcellbyvalue(value, worksheet_to_search):
#Поиск ячейки по значению
    print('Поиск ячейки ' + str(value) + ' в таблице ' + str(worksheet_to_search.name))
    for row in range(worksheet_to_search.nrows):
        for col in range(worksheet_to_search.ncols):
            cell = worksheet_to_search.cell(row,col)
            #print (str(cell.value))
            if str(value) == str(cell.value) or str(value) in str(cell.value):
                #print (str(cell.value))
                return(row,col)


def findcellbyvalue_onecolumn(value, worksheet_to_search,col):
#Бинарный поиск ячейки по значению
	p = 1
	r = worksheet_to_search.nrows
	while (p <= r):
		q = (r + p)//2
		cell = worksheet_to_search.cell(q,0)
		if (str(value) == str(cell.value)) or (str(value) + '=' == str(cell.value)):
			return(q,col)
		elif str(value) > str(cell.value):
			p = q + 1
		else:
			r = q - 1
	return (None,col)
 
os.chdir(working_dir)
#Открытие исходной спецификации
try:
    old_workbook = xlrd.open_workbook(working_spec_filename)
except:
    print('Ошибка открытия исходной спецификации')
old_worksheet = old_workbook.sheet_by_index(0)

#Открытие CPL и таблицы с категориями в нём
if (not os.path.isfile('Cisco Product List Sorted.xls')):
	insertion_sort_cpl()
	
try:
    cpl_workbook = xlrd.open_workbook('Cisco Product List Sorted.xls')
except IOError as e:
    print ("I/O error({0}): {1}".format(e.errno, e.strerror))
    #print('Ошибка открытия Cisco Product List')
cpl_worksheet = cpl_workbook.sheet_by_name('Cisco Product List Russia')

#Найти начало колонки с партномером
pn_cell = findcellbyvalue('Part Number',old_worksheet)
print('Начало колонки с партномерами: ' + str(pn_cell))
#Найти начало колонки с описанием
descr_cell = findcellbyvalue('Description',old_worksheet)

#Найти начало колонки с GPL за единицу
gpl_cell = findcellbyvalue('Unit List Price',old_worksheet)

#Найти начало колонки с количеством 
qty_cell = findcellbyvalue('Qty',old_worksheet)

new_workbook = openpyxl.Workbook()
new_worksheet = new_workbook.active

#Заголовки колонок
new_worksheet['A1'] = '№ п.п.'
new_worksheet['B1'] = 'Партномер'
new_worksheet['C1'] = 'Описание'
new_worksheet['D1'] = 'GPL за ед.'
new_worksheet['E1'] = 'Кол-во'
new_worksheet['F1'] = 'GPL за поз.'
new_worksheet['J1'] = 'Категория C'
new_worksheet['K1'] = 'Категория Rn'
#Жирный шрифт
for i in range (1, 20):
    new_worksheet.cell(row=1, column=i).font = openpyxl.styles.Font(bold=True, name='Calibri')
    

#Сделать правильную ширину колонок
#

i = pn_cell[0]+1
j = 2
pn_regexp = re.compile('^[A-Z]+[A-Z0-9\-\=\+\/\.]+$')


while pn_regexp.match(old_worksheet.cell(i, pn_cell[1]).value) != None or pn_regexp.match(old_worksheet.cell(i+1, pn_cell[1]).value) != None:
#Если не партномер, пропустить цикл
    if pn_regexp.match(old_worksheet.cell(i, pn_cell[1]).value) == None:
        i += 1
        continue															
#Порядковый номер строки
    new_worksheet.cell(row=j,column=1).value = j-1
#Партномер
    new_worksheet.cell(row=j,column=2).value = old_worksheet.cell(i, pn_cell[1]).value
#Описание
    new_worksheet.cell(row=j,column=3).value = old_worksheet.cell(i, descr_cell[1]).value
#GPL за единицу
    new_worksheet.cell(row=j,column=4).value = old_worksheet.cell(i, gpl_cell[1]).value
    new_worksheet.cell(row=j,column=4).number_format = '0.00$'
#Количество
    new_worksheet.cell(row=j,column=5).value = old_worksheet.cell(i, qty_cell[1]).value
#GPL за позицию
    new_worksheet.cell(row=j,column=6).value = '='+ xlrd.cellname(j-1, gpl_cell[1]-1) + '*' + xlrd.cellname(j-1, qty_cell[1]-1)#old_worksheet.cell(i, gpl_cell[1]).value*old_worksheet.cell(i, qty_cell[1]).value
    new_worksheet.cell(row=j,column=6).number_format = '0.00$'

    print('Партномер: ' + old_worksheet.cell(i, pn_cell[1]).value)
#Категории Rn и C
  
#Если позиция сервисная, категория не проверяется
    if old_worksheet.cell(i, pn_cell[1]).value.startswith('CON-'):
        new_worksheet.cell(row=j,column=10).value = 'Сервисная позиция. Категория не применяется'
        new_worksheet.cell(row=j,column=11).value = 'Сервисная позиция. Категория не применяется'
        print('Сервисная позиция. Категория не применяется')
        j += 1
    else:
#Если не сервисная, ищем в Product List и подставляем значения C и Rn
    
        try:
            c_cat_cell = findcellbyvalue_onecolumn(old_worksheet.cell(i, pn_cell[1]).value, cpl_worksheet,0)
            print('Найден в CPL, ячейка ' + str(c_cat_cell))
            new_worksheet.cell(row=j,column=10).value = cpl_worksheet.cell(c_cat_cell[0], c_cat_cell[1]+1).value
            new_worksheet.cell(row=j,column=11).value = cpl_worksheet.cell(c_cat_cell[0], c_cat_cell[1]+10).value
#C3/C4 помечаем красным и полужирным
            if new_worksheet.cell(row=j,column=10).value == 'C3' or new_worksheet.cell(row=j,column=10).value == 'C4':
                new_worksheet.cell(row=j, column=10).font = openpyxl.styles.Font(bold=True, color=openpyxl.styles.colors.RED)
        except Exception as e:
            new_worksheet.cell(row=j,column=10).value = 'Нет категории'
            #print (e)
        
        j += 1
    i += 1

#Установить ширину колонок
new_worksheet.column_dimensions['A'].width = 8
new_worksheet.column_dimensions['B'].width = 22
new_worksheet.column_dimensions['C'].width = 61
new_worksheet.column_dimensions['D'].width = 12
new_worksheet.column_dimensions['E'].width = 9
new_worksheet.column_dimensions['F'].width = 12
new_worksheet.column_dimensions['J'].width = 42
new_worksheet.column_dimensions['K'].width = 42

#Сохранить новый файл
print('*******************')
print('Запись в файл ' + resulting_file)
try:
    new_workbook.save(resulting_file)
    print('Успешно')
except:
    print('Ошибка записи в файл')

    

